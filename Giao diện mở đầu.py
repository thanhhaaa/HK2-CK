from tkinter import *
from tkinter import ttk, messagebox
from datetime import datetime
from openpyxl import load_workbook
import openpyxl
from tkinter import ttk

now = datetime.now()

window = Tk()
window.title("MANAGE DAILY CALORIE INTAKE ")
window.geometry("650x680+320+50")
window.resizable(width=False, height=False)
window.config(bg="#8EB5B4")

mynotebook = ttk.Notebook(window)
mynotebook.pack(fill=BOTH)


def tab():
    myframe = Frame(mynotebook, bg="#8EB5B4")
    myframe.pack(fill=BOTH, expand=True)
    mynotebook.add(myframe, text="Menu")
    mylabel = Label(myframe, text='Choose feature', font="Arial 20 bold")
    mylabel.pack(fill=BOTH, expand=True)
    bttn1 = Button(myframe, text="Kcal Calculator", font="Arial 15", command=tab1, width=25)
    bttn1.pack(padx=50,pady=15)
    bttn2 = Button(myframe, text="TDEE Calculator", font="Arial 15",command=tab2, width=25)
    bttn2.pack(padx=50, pady=15)
    bttn3 = Button(myframe, text="Note", font="Arial 15", command=tab3, width=25)
    bttn3.pack(padx=50, pady=15)


def tab1():
    myframe = Frame(mynotebook, bg="#8EB5B4")
    myframe.place(x=50, y=50)
    mynotebook.add(myframe, text="Kcal Calculator")
    mynotebook.select(myframe)

    framecbb = Frame(myframe)
    framecbb.pack(side=LEFT)
    mylabel = Label(framecbb, text='ENTER FOOD')
    mylabel.pack()

    mytext = Text(framecbb, width=28, height=-5, font=("Ariel", 10))
    mytext.pack()

    weight = Label(framecbb, text="FOOD INTAKE (gam)")
    weight.pack()

    value = StringVar(value=0)
    ws = ttk.Spinbox(framecbb, from_=1, to=1000, textvariable=value, wrap=True, width=30)
    ws.pack(anchor=CENTER)

    framemt = Frame(myframe)
    framemt.pack()

    mt = ttk.Treeview(framemt, columns=("1", "2", "3"), show='headings')

    mt.column("#0", width=0)
    mt.column("1", width=150)
    mt.column("2", width=90)
    mt.column("3", width=150)


    mt.heading("#0")
    mt.heading("1", text="Foods")
    mt.heading("2", text="Quantity")
    mt.heading("3", text="Kcal")

    mt.pack(anchor=NE)


    list = []

    def delete():
        select_item = mt.selection()[0]
        mt.delete(select_item)

    path = "C:\\Users\\Lac Thu Thuy\\Downloads\\calo.xlsx" #C:\\Users\\data\\Downloads\\calo.xlsx (do ổ đĩa t khác)
    wb_obj = openpyxl.load_workbook(path)

    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row

    listfood = []
    listkcal = []

    for i in range(1, m_row + 1):
        cell1_obj = sheet_obj.cell(row=i, column=1)
        listfood.append(cell1_obj.value)

    print(listfood)
    for i in range(1, m_row + 1):
        cell2_obj = sheet_obj.cell(row=i, column=2)
        listkcal.append(cell2_obj.value)


    def cal():
        for i in range(len(listfood)):
            if mytext.get("1.0", "end-1c") == listfood[i]:
                a = float(ws.get()) * float(listkcal[i])
                print("2", a)
                return a


    def getdata():
        string = mytext.get("1.0", "end-1c") + " " + ws.get() + " " + str(cal())
        list.append(string)
        mt.insert(parent='', index=END, values=list[len(list) - 1])


    def cal():
        for i in range(len(listfood)):
            if mytext.get("1.0", "end-1c") == listfood[i]:
                a = float(ws.get()) * float(listkcal[i])
                print("2", a)
                return a



    def getdata():
        string = mytext.get("1.0", "end-1c") + " " + ws.get() + " " + str(cal())
        list.append(string)
        mt.insert(parent='', index=END, values=list[len(list) - 1])

    mybtt1 = Button(framecbb, text='Save', command=getdata)
    mybtt1.pack()
    mybttn2 = Button(framemt, text="Delete", command=delete)
    mybttn2.pack()
    mybttn3 = Button(framemt, text="Calculate", command=cal)
    mybttn3.pack()

    def exit():
        op = messagebox.askyesno("Confirm", "Do you want to leave this page?")
        if op == True:
            myframe.destroy()

    but1 = Button(myframe, text='Quit', command=exit)
    but1.pack(side=RIGHT, anchor=S)


def tab2():
    myframe = Frame(mynotebook, bg="#8EB5B4" )
    myframe.place(x=350, y=600)
    mynotebook.add(myframe, text="TDEE Calculator")
    mynotebook.select(myframe)

    class TDEE_Class:
        def __init__(self, myframe, mynotebook):
            self.myframe = myframe
            self.mynotebook = mynotebook
            lb = Label(myframe, text="Calories per day", font=("Arial bold", 25),bg="#8EB5B4")
            lb.pack(side=TOP, fill=X)

            self.var_height = StringVar()
            self.var_weight = StringVar()
            self.var_age = StringVar()
            self.var_workout = StringVar()
            self.var_gender = StringVar()
            self.var_yourTDEE = StringVar()

            lbl_age = Label(myframe, text="Age", font=("Arial", 20), bg="#8EB5B4")
            lbl_age.place(x=50, y=100)

            txt_age = Entry(myframe, textvariable=self.var_age, font=("Arial", 15), bg="white")
            txt_age.place(x=250,y=100,width=300,height=40)

            lbl_gender = Label(myframe, text="Gender", font=("Arial", 20), bg="#8EB5B4")
            lbl_gender.place(x=50, y=170)

            gender = ["Male", "Female"]
            rbt = ttk.Combobox(myframe, values=gender, textvariable=self.var_gender, font="Arial 15")
            rbt.place(x=250, y=170,width=300,height=40)


            lbl_height = Label(myframe, text="Height", font=("Arial", 20), bg="#8EB5B4")
            lbl_height.place(x=50,y=240)
            txt_height = Entry(myframe, textvariable=self.var_height, font=("Arial", 15), bg="white")
            txt_height.place(x=250, y=240, width=300, height=40)

            lbl_weight = Label(myframe, text="Weight", font=("Arial", 20), bg="#8EB5B4")
            lbl_weight.place(x=50,y=310)
            txt_weight = Entry(myframe, textvariable=self.var_weight, font=("Arial", 15), bg="white")
            txt_weight.place(x=250, y=310, width=300, height=40)

            lbl_workout = Label(myframe, text="Work out", font=("Arial", 20), bg="#8EB5B4")
            lbl_workout.place(x=50,y=380)
            workout = ["Inactive", "Light exercise (1-3 days/week)", "Average exercise (3-5 days/week)",
                       "Exercise a lot (6-7 days/week)", "High intensity exercise (everyday)"]
            Cmb = ttk.Combobox(myframe, values=workout, textvariable=self.var_workout, font="Arial 15")
            Cmb.place(x=250,y=380,width=300,height=40)


            btn_TDEE = Button(myframe, text=" Calculate TDEE", bg="#DCF8C6", command=self.TDEE, font=("Arial", 15))
            btn_TDEE.place(x=250,y=430,width=160,height=50)
            btn_clear = Button(myframe, text="Clear", bg="#BBBBBB", command=self.clear, font=("Arial", 15))
            btn_clear.place(x=420,y=430,width=130,height=50)

            lbl_your_TDEE = Label(myframe, text="Your TDEE", font=("Arial", 20), bg="#8EB5B4")
            lbl_your_TDEE.place(x=50, y=530)
            txt_your_TDEE = Entry(myframe, textvariable=self.var_yourTDEE, font=("Arial", 20), bg="white",
                                  fg="#cc2900")
            txt_your_TDEE.place(x=250, y=530, width=300, height=40)
            btn_exit = Button(myframe, text="Exit", fg="white", bg="#3B5988", command=self.exit,
                              font=("Arial", 15))
            btn_exit.place(x=300, y=590, width=130, height=50)

            note = Label(myframe, text="TDEE indicates the number of calories you need to eat in a day", font="Arial 13")
            note.pack(side=BOTTOM)

        def clear(self):
            self.var_age.set("")
            self.var_gender.set("")
            self.var_height.set("")
            self.var_weight.set("")
            self.var_workout.set("")
            self.var_yourTDEE.set("")

        def TDEE(self):
            self.g = (self.var_gender.get())
            if self.var_age.get() == "" or self.var_gender.get() == "" or self.var_height.get() == "" or self.var_weight.get() == "" or self.var_workout.get() == "":
                messagebox.showerror("Error", "Please enter full information")
            if self.g == "Male":
                self.a = int(self.var_age.get())
                self.h = float(self.var_height.get())
                self.w = float(self.var_weight.get())
                self.wk = (self.var_workout.get())
                self.number = round((self.w * 10) * 2.2046 + (self.h) * 6.25 * 0.39370 - (self.a * 5) + 5)
                if self.wk == "Inactive":
                    self.TDEE = round((self.number * 1.2), 2)
                if self.wk == "Light exercise (1-3 days/week)":
                    self.TDEE = round((self.number * 1.375), 2)
                if self.wk == "Average exercise (3-5 days/week)":
                    self.TDEE = round((self.number * 1.55), 2)
                if self.wk == "Exercise a lot (6-7 days/week)":
                    self.TDEE = round((self.number * 1.725), 2)
                if self.wk == "High intensity exercise (everyday)":
                    self.TDEE = round((self.number * 1.9), 2)
                self.var_yourTDEE.set(self.TDEE)
            elif self.g == "Female":
                self.a = int(self.var_age.get())
                self.h = float(self.var_height.get())
                self.w = float(self.var_weight.get())
                self.wk = (self.var_workout.get())
                self.number = round((self.w * 10) * 2.2 + (self.h) * 6.25 * 0.39 - (self.a * 5) - 161, 1)
                if self.wk == "Inactive":
                    self.TDEE = round((self.number * 1.2), 2)
                if self.wk == "Light exercise (1-3 days/week)":
                    self.TDEE = round((self.number * 1.375), 2)
                if self.wk == "Average exercise (3-5 days/week)":
                    self.TDEE = round((self.number * 1.55), 2)
                if self.wk == "Exercise a lot (6-7 days/week)":
                    self.TDEE = round((self.number * 1.725), 2)
                if self.wk == "High intensity exercise (everyday)":
                    self.TDEE = round((self.number * 1.9), 2)
                self.var_yourTDEE.set(self.TDEE)

        def exit(self):
            op = messagebox.askyesno("Confirm", "Do you want to leave this page?")
            if op == True:
                self.myframe.destroy()
    TDEE_Class(myframe, mynotebook)


def tab3():
    myframe = ttk.Frame(mynotebook)
    myframe.pack()
    mynotebook.add(myframe, text="Note")
    mynotebook.select(myframe)

    class Tab:
        def __init__(self, myframe, mynotebook):
            self.myframe = myframe
            self.mynotebook = mynotebook
            self.tab()

        def ghiFile(self, path, content):  # Lưu file
            f = open(path, "a", encoding="utf-8")
            f.write('-----' + now.strftime("%d") + "/" + now.strftime("%m") + "/" + now.strftime("%Y") + '-----' + "\n")
            f.write(content + "\n" + "\n")
            f.close()

        path = "run.txt"

        def tab(self):
            def save():
                inputValue = textbox.get("1.0", "end-1c")  # Lấy giá trị từ textbox
                self.ghiFile("run.txt", inputValue)
            textbox = Text(myframe, height=15)
            textbox.pack(fill=BOTH)
            but3 = Button(myframe, text='Quit', command=self.exit, relief=GROOVE)
            but3.pack(side=RIGHT)
            but1 = Button(myframe, text="Save", command=save, relief=GROOVE)
            but1.pack(side=RIGHT)


        def exit(self):
            op = messagebox.askyesno("Confirm", "Do you want to leave this page?")
            if op == True:
                self.myframe.destroy()

    Tab(myframe, mynotebook)


abu = Text()
tab()
window.mainloop()
